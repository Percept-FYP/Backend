
from django.shortcuts import render
from openpyxl.cell import cell
from api.recognize_faces_image import func
from django.http import JsonResponse
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, DjangoMultiPartParser
from rest_framework.decorators import api_view
from api.serializers import AttendanceSerializer, ClassSerializer, RegisterSerializer, SubjectsSerializer, Time_tableSerializer
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from api.models import *
import PIL.Image
import openpyxl
import cv2
import datetime
from django.core.files.base import File
import os
from pathlib import Path
from rest_framework import status
import xlsxwriter
from datetime import datetime
BASE_DIR = Path(__file__).resolve().parent.parent


@api_view(['POST'])
def create(request):
    usn  = "1JT17CS0"
    for  i in range(1,60):
        usn_ = usn+str(i)
        student = Student.objects.create(usn=usn)
    return Response("created ")


@api_view(['POST'])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    print(serializer)
    if serializer.is_valid(raise_exception=True):
        print("were hered")

        serializer.save()
        print(serializer.data)
        return Response(serializer.data)
    else:
        return Response("not created")


@api_view(['POST'])
def user_details(request):
    id = request.data['user_id']
    first_name = request.data['first_name']
    last_name = request.data['last_name']
    role = request.data['role']
    profile_image = request.data['profile_image']
    phone = request.data['phone']
    try:
        designation = request.data["designation"]
    except:
        designation = "null"
    try:
        usn = request.data['usn']
        sem = request.data['sem']
        branch = request.data['branch']
    except:
        usn = "null"
    user = User.objects.get(id=id)
    user.first_name = first_name
    user.last_name = last_name
    user.role = role
    user.phone = phone
    user.image = profile_image
    user.save()
    if user.role == "teacher":
        Teacher = teacher.objects.create(user=user, designation=designation)
        return Response("you are a teacher now")
    elif user.role == "student" and usn != "null":

        student = Student.objects.create(
            user=user, usn=usn, name=first_name, academic_info=Academic_info.objects.get(branch=branch,semester=sem))
        student.academic_info.semester = sem

        student.save()
        return Response("you are a student")
    else:
        return Response("no role created")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def post(request):
    data = request.data
    print("data",data)
    image1 = ""
    image2 = ""
    if data["image1"]:
        image1 = data["image1"]
    if data["image2"]:
        image2 = data["image2"]

    CLASS = Class.objects.create(subject=Subject.objects.get(
        subject_code=data["class_name"]), image=data["image"], image1=image1, image2=image2)
    id = CLASS.id
    print(CLASS.image)
    serializer = ClassSerializer(CLASS)
    if serializer:
        print("valid")
        list_all = []
        for e in Student.objects.all().order_by("usn"):
            list_all.append(e.usn)
        images = []
        images.append(CLASS.image)
        images.append(CLASS.image1)
        images.append(CLASS.image2)

        file = CLASS.subject.attendance_file
        print(file)
        present_list = []
        for img in images:
            present_list = present_list + func(img)
        print("present_list", present_list)
        abs_list = list(set(list_all) - set(present_list)) + \
            list(set(present_list) - set(list_all))
        class_na = data["class_name"]

        present_list = list(set(present_list))
        try:
            present_list.remove('Unknown')
        except:
            pass
        try:
            abs_list.remove('Unknown')
        except:
            pass
        for usn in present_list:
            print(usn)
            Student1,create =Student.objects.get_or_create(usn=usn)
            attend = attendance.objects.create(
                student=Student1, attend="present", Class=Class.objects.get(id=id))
        for usn in abs_list:
            print(usn)
            Student1,create =Student.objects.get_or_create(usn=usn)
            attend = attendance.objects.create(
                student=Student1, attend="absent", Class=Class.objects.get(id=id))
        attendances = attendance.objects.filter(
            Class=id).order_by("student")
        pth = os.path.join(BASE_DIR, "static\\media\\")+str(file)
        p = pth.replace("/", "\\")

        filename = str(CLASS.subject)

        exl = p.replace("somefile1", filename)
        
        partitioned_string = exl.rpartition('_')
        print(partitioned_string)
        try:
            workbook1 = openpyxl.load_workbook(partitioned_string[0]+'.xlsx')
        
        except:
            workbook1 = openpyxl.Workbook()
            workbook1.save(partitioned_string[0]+'.xlsx')
       
        # get the number of columns filled
       

        # Get the currennt date
        today = datetime.date.today().strftime("%d-%m-%Y")
        wbkName = (partitioned_string[0]+'.xlsx')
        wbk = openpyxl.load_workbook(wbkName)
        worksheet1 = wbk['Sheet']
        ncol = worksheet1.max_column

        row = 0
        column = 0
        # Loop through the usn and write to the column 1
        for wks in wbk.worksheets:
            i = 0
            wks.cell(row=i+1, column=1).value = "USN"
            while i < len(list_all):
                wks.cell(row=i+2, column=1).value = list_all[i]
                i += 1

        # Check the result names and assign the attendance in the excel sheet
        for wks in wbk.worksheets:
            i = 0
            wks.cell(row=i+1, column=ncol+1).value = today
            while i < len(list_all):
                if list_all[i] in present_list:
                    wks.cell(row=i+2, column=ncol+1).value = "Present"
                else:
                    wks.cell(row=i+2, column=ncol+1).value = "Absent"
                i += 1

        # Save the workbook
        wbk.save(wbkName)
        wbk.close

    print(abs_list)
    print(present_list)
    CS = ClassSerializer(instance=CLASS)
    print(CS.data)
    serializer = AttendanceSerializer(attendances, many=True)
    print(serializer.data)
    return Response(CS.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update(request, cl):
    # attendances = attendance.objects.filter(Class=cl).delete()
    attendancse = request.data['attendance_set']
    clas = Class.objects.get(id=cl)
    file = clas.subject.attendance_file
    # serializer = ClassSerializer(instance=clas, data=request.data)
    for atd in attendancse:
        id = atd['id']
        isinst = attendance.objects.get(id=id)
        serializer2 = AttendanceSerializer(instance=isinst, data=atd)
        if serializer2.is_valid():
            serializer2.save()
    if serializer2.is_valid():
        serializer2.save()
        present_list = []
        list_all = []
        attends = attendance.objects.filter(Class_id=cl, attend="present")
        for i in attends:
            present_list.append(i.student.usn)
        for e in Student.objects.all().order_by("usn"):
            list_all.append(e.usn)
        print(list_all)
        pth = os.path.join(BASE_DIR, "static\\media\\")+str(file)
        p = pth.replace("/", "\\")

        filename = str(clas.subject)

        exl = p.replace("somefile1", filename)
        print(exl)
        partitioned_string = exl.rpartition('_')

        workbook1 = openpyxl.load_workbook(partitioned_string[0]+'.xlsx')
        worksheet1 = workbook1['Sheet']
        # get the number of columns filled
        ncol = worksheet1.max_column

        row = 0
        column = 0

        # Get the currennt date
        today = datetime.date.today().strftime("%d-%m-%Y")
        wbkName = partitioned_string[0]+'.xlsx'
        wbk = openpyxl.load_workbook(wbkName)

        # Loop through the usn and write to the column 1
        for wks in wbk.worksheets:
            i = 0
            wks.cell(row=i+1, column=1).value = "USN"
            while i < len(list_all):
                wks.cell(row=i+2, column=1).value = list_all[i]
                i += 1

        # Check the result names and assign the attendance in the excel sheet
        for wks in wbk.worksheets:
            i = 0
            wks.cell(row=i+1, column=ncol).value = today
            while i < len(list_all):
                if list_all[i] in present_list:
                    wks.cell(row=i+2, column=ncol).value = "Present"
                else:
                    wks.cell(row=i+2, column=ncol).value = "Absent"
                i += 1

        # Save the workbook
        wbk.save(wbkName)
        wbk.close

    return Response("Sucessgully Updated", status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def subject_create(request):
    Teacher = teacher.objects.get(user=request.user)
    subject_name = request.data['subject_name']
    subject_code = request.data['subject_code']
    sem = request.data['sem']
    branch = request.data['branch']
    scheme = request.data['scheme']
    academic_info,created=Academic_info.objects.get_or_create(branch=branch,semester=sem,scheme=scheme)
    subject = Subject.objects.create(teacher=Teacher,
                                     subject_name=subject_name, subject_code=subject_code, academic_info=academic_info)
    
    

    subject.save()
    return Response("subject added")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subjects(request):
    if request.user.role == "teacher":
        print('teacher')
        subjects = Subject.objects.filter(
            teacher=teacher.objects.get(user=request.user))
        print(subjects)
        serializer = SubjectsSerializer(subjects, many=True)
        return Response(serializer.data)
    elif request.user.role == "student":
        print('students')
        student = Student.objects.get(user=request.user)
        academic_info = student.academic_info
        subjects = Subject.objects.filter(
            academic_info=academic_info)
        print(subjects)
        serializer = SubjectsSerializer(subjects, many=True)
        return Response(serializer.data)
    else:
        print('no role')
        return Response("no role set")


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def attendance_details(request, cl):

    if request.method == 'GET':
        user = request.user
        student = Student.objects.get(user=user)
        Clasess = Class.objects.filter(
            subject=Subject.objects.get(subject_code=cl))
        attendances = []
        for CLASS in Clasess:
            Attendance = attendance.objects.filter(
                student=student, Class=CLASS)
            serializer = AttendanceSerializer(Attendance, many=True)
            attendances = attendances + serializer.data
        print(attendances)
        return Response(attendances)

    elif request.method == 'POST':
        usn = request.data['usn']
        student = Student.objects.get(usn=usn)
        Clasess = Class.objects.filter(
            subject=Subject.objects.get(subject_code=cl))
        attendances = []
        for CLASS in Clasess:
            Attendance = attendance.objects.filter(
                student=student, Class=CLASS)
            serializer = AttendanceSerializer(Attendance, many=True)
            attendances = attendances + serializer.data
        print(attendances)
        return Response(attendances, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def attendance_overview(request):
    user = request.user
    student = Student.objects.get(user=user)
    Academic_info = student.academic_info
    subjects = Subject.objects.filter(academic_info=Academic_info)
    result_set = []
    for subject in subjects:
        present_count = 0
        absent_count = 0
        all = 0
        Clasess = Class.objects.filter(
            subject=subject)
        for CLASS in Clasess:
            print(CLASS)
            all = all + 1

            present_count = present_count + attendance.objects.filter(
                student=student, Class=CLASS, attend="present").count()
            print(present_count)
            absent_count = absent_count + attendance.objects.filter(
                student=student, Class=CLASS, attend="absent").count()
            print(absent_count)

        print("subject:", str(subject), "present:", present_count,
              "absent:", absent_count)
        print("total", present_count+absent_count)
        if present_count > 0:
            # have to change!
            percentage = (present_count/(present_count+absent_count))*100
        else:
            percentage = 0
        result = {
            "subject": str(subject),
            "classes attended":  present_count,
            "classes taken": all,
            "%": round(percentage, 2)
        }
        result_set.append(result)
    attendances_overview = []

    return Response(result_set)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def attendance_filter(request):
    date = request.data['date']
    Sub = request.data['subject']
    user = request.user
    student = Student.objects.get(user=user)
    Clasess = Class.objects.filter(
        subject=Subject.objects.get(subject_code=Sub), time__date=date)
    attendances = []
    for CLASS in Clasess:
        Attendance = attendance.objects.filter(
            student=student, Class=CLASS)
        serializer = AttendanceSerializer(Attendance, many=True)
        attendances = attendances + serializer.data
    print(attendances)
    return Response(attendances)


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def attendance_filter(request):
#     user = request.user
#     student = Student.objects.get(user=user)
#     Academic_info = student.Academic
#     Clasess = Class.objects.filter(
#         subject=Subject.objects.get(subject_code=Sub), time__date=date)
#     attendances = []
#     for CLASS in Clasess:
#         Attendance = attendance.objects.filter(
#             student=student, Class=CLASS)
#         serializer = AttendanceSerializer(Attendance, many=True)
#         attendances = attendances + serializer.data
#     print(attendances)
#     return Response(attendances)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def time_table(request):
    user = request.user
    if user.role == "student":
        student = Student.objects.get(user=user)
        academic_info = student.academic_info.id
        print("academic", academic_info)
        subjects = Subject.objects.filter(academic_info=academic_info)
        time_table = []
        for subject in subjects:
            print(subject)
            time_tables = Time_table.objects.filter(subject=subject)
            serializer = Time_tableSerializer(time_tables, many=True)
            print(serializer.data)
            time_table = time_table + serializer.data
        print(time_table)
        return Response(time_table)

    else:
        id = user.id
        Teacher = teacher.objects.get(user=id)
        subjects = Subject.objects.filter(teacher=Teacher)
        time_table = []
        for subject in subjects:
            print(subject)
            time_tables = Time_table.objects.filter(subject=subject)
            serializer = Time_tableSerializer(time_tables, many=True)
            print(serializer.data)
            time_table = time_table + serializer.data
        print(time_table)
        return Response(time_table)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def recent_class(request):
    id = request.user.id
    Teacher = teacher.objects.get(user=id)
    print("teacher", Teacher)
    recent_class = Class.objects.filter(subject__teacher__user=id).last()
    seralizer = ClassSerializer(recent_class)
    print("recent_Class", recent_class)
    present_count = attendance.objects.filter(
        Class=recent_class, attend="present").count()
    print("p", present_count)

    absent_count = attendance.objects.filter(
        Class=recent_class, attend="absent").count()
    print("a", absent_count)
    try: 
        percentage = (present_count/(present_count+absent_count))*100
    except:
        percentage = 0
    data = {
        "percentage":round(percentage,2),
        "recent_class":seralizer.data
    }
    print("percentage", round(percentage,2))
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def today_schedule(request):
    print(request.user)
    day = request.data['day']
    if request.user.role == "teacher":
        print("teacher")
        id = request.user.id
        schedule = Time_table.objects.filter(day=day, subject__teacher__user=id)
        serializer = Time_tableSerializer(schedule, many=True)
        return Response(serializer.data)
    elif request.user.role == "student":
        print("student")
        student = Student.objects.get(user=request.user)
        schedule = Time_table.objects.filter(day=day, subject__academic_info=student.academic_info) 
        serializer = Time_tableSerializer(schedule, many=True)
        return Response(serializer.data)
    else:
          return Response("hod")
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def xl_filter(request):
    #path = "C:\\Users\\supp\\Desktop\\Python\\Example1.xlsx"
    #olddd - C:/Users/supp/Desktop/Python/Example1.xlsx
    print(request.data)
    subject_code  = request.data['subject_code']
    from_date = request.data['from_date']
    to_date = request.data['to_date']
    pth = os.path.join(BASE_DIR, "static\\media\\records\\")+str(subject_code)
    wb_obj = openpyxl.load_workbook(pth+".xlsx")
    sheet = wb_obj.active
    ncol = sheet.max_column 
    nrow=sheet.max_row
    
    # cell_obj = sheet.cell(row = 5, column = 6)
 

    fromm=0
    too=0
    flag1=0
    flag2=0
    # print("cel_obj",(str(cell_obj.value).split(' ')[0]))
    for i in range(1,ncol+1):
        cell_obj = sheet.cell(row = 1, column = i)
        print("cel_obj",(str(cell_obj.value).split(' ')[0]))
        print("from_date",from_date)
        if from_date==(str(cell_obj.value).split(' ')[0]):
            fromm=i
            flag1=1
            
        if to_date==(str(cell_obj.value).split(' ')[0]):
            too=i
            flag2=1

    if flag1==0:
        print("from date not present")
    elif flag2==0:
        print("to date not present")
    elif from_date>to_date:
        print("from>to error")
    else:    
        usns=[]
        arr=[]
        for i in range(2,nrow+1):
            cell_obj = sheet.cell(row = i, column = 1)
            usns.append(cell_obj.value)
        for i in range(1, nrow+1):
            col=[]
            for j in range(fromm,too+1):
                cell_obj = sheet.cell(row = i, column = j)
                if cell_obj.value != "Absent" and cell_obj.value != "Present":
                    temp = (str(cell_obj.value).split(' ')[0])
                    t2 = temp.split('-')
                    t3 = t2[2] + '-' + t2[1] + '-'+ t2[0]
                    col.append(t3)
                else:
                    col.append(cell_obj.value)
            arr.append(col)
        workbook = xlsxwriter.Workbook(pth+'_temp.xlsx')
        worksheet = workbook.add_worksheet()
        row = 0
        col = 0
        worksheet.write(0, 0,'USN')
        for i in range(0,nrow):
            if(i!=nrow-1):
                worksheet.write(i+1, 0,usns[i])
            for j in range(0,(too-fromm)+1):
                worksheet.write(i, j+1,arr[i][j])

        workbook.close()
    url = 'media/records/'+subject_code + "_temp.xlsx" 
    return Response(url)